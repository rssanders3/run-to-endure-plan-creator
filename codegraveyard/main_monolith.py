import gspread
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
import csv
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
import textwrap

DAY_NUM_WEEK_MAP = {
    1: "Monday",
    2: "Tuesday",
    3: "Wednesday",
    4: "Thursday",
    5: "Friday",
    6: "Saturday",
    7: "Sunday"
}

# Define the scope and authenticate
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("PATH TO CONFIG", scope)
client = gspread.authorize(creds)

SPREADSHEET_URL = "URL TO YOUR SPREADSHEET"
# Open your Google Sheet by name or by URL
spreadsheet = client.open_by_url(SPREADSHEET_URL)
worksheet = spreadsheet.worksheet("Planning")

# Get the value of cell H1 - AKA Swim Pace
swim_pace = worksheet.acell("H1").value
print(f"Swim Pace: {swim_pace}")
# Get the value of cell M1 - AKA Bike Speed
bike_speed = worksheet.acell("M1").value
print(f"Bike Speed: {bike_speed}")
# Get the value of cell R1 - AKA Run Pace
run_pace = worksheet.acell("R1").value
print(f"Run Pace: {run_pace}")

# Get values in the range D5:AL74
data = worksheet.get("D5:AL14")

training_plan_master = {}
# Process the data: group every 5 rows as one week and every 5 columns as one day
week_counter = 0
for week_start in range(0, len(data), 5):  # Iterate over weeks (5 rows at a time)
    week_counter += 1
    print("week num: ", week_counter)
    week_data = data[week_start:week_start + 5]  # Get 5 rows for the week
    week_output_data = {}
    day_num = 0
    for day_start in range(0, len(week_data[0]), 5):  # Iterate over days (5 columns at a time)
        day_num += 1
        print("day num: ", day_num)
        day_values = [week_row[day_start:day_start + 5] for week_row in week_data]  # Get 5 columns for the day
        print(f"Day values: {day_values}")
    
        swim_data = day_values[0]
        bike_data = day_values[1]
        run_data = day_values[2]
        strength_data = day_values[3]

        discipline = ""
        workout_type = ""
        summary = ""
        if swim_data[0] != "":
            discipline += "Swim,"
            workout_type += "Technique Swim,"
            summary += textwrap.dedent(f"""
            Stretch
                                       
            Main Set:
            {swim_data[1]}-km @ Z2 (RPE: {swim_data[3]}),
            """)
        if bike_data[0] != "":          
            discipline += "Bike,"
            workout_type += "Easy Ride,"
            summary += textwrap.dedent(f"""
            Stretch
                                       
            Main Set:
            {bike_data[1]}-km @ Z2 (RPE: {bike_data[3]}),
            """)
        if run_data[0] != "":
            discipline += "Run,"
            workout_type += "Easy Run,"
            # summary += "Run " + run_data[0] + " Minutes, "
            summary += textwrap.dedent(f"""
            Stretch
                                       
            Main Set:
            {run_data[1]}-km @ Z2 (RPE: {run_data[3]}), 
            """)
        if strength_data[0] != "":
            discipline += "Workout,"
            workout_type += "Stretch & Strength,"
            summary += "See the Stretch Routine & Strength Training section below,"
        if swim_data[0] == "" and bike_data[0] == "" and run_data[0] == "" and strength_data[0] == "":
            discipline = "Rest Day"
            workout_type = ""
            summary += ""

        week_output_data[DAY_NUM_WEEK_MAP[day_num]] = {
            "Discipline": discipline.strip(","),
            "WorkoutType": workout_type.strip(","),
            "Summary": summary.strip(",")
        }

    training_plan_master[f"Week {week_counter}"] = week_output_data

# print(training_plan_master)

output_rows = []
# Flatten and write training_plan_master to a CSV file
for week, days in training_plan_master.items():
    print(f"{week}:")
    output_rows.append(["Week", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"])
    week_disciplin_row = [week]
    week_workout_type_row = [week]
    week_summary_row = [week]
    for day, details in days.items():
        week_disciplin_row.append(details["Discipline"])
        week_workout_type_row.append(details["WorkoutType"])
        week_summary_row.append(details["Summary"])
    output_rows.append(week_disciplin_row)
    output_rows.append(week_workout_type_row)
    output_rows.append(week_summary_row)
    output_rows.append(["", "", "", "", "", "", "", ""])

with open("training_plan_master.csv", mode="w", newline="") as file:
    writer = csv.writer(file)
    writer.writerows(output_rows)
print("Training Plan saved to `training_plan_master.csv`")

# # Convert the above process to a DataFrame
# summary_df = pd.DataFrame(output_rows)
# # output the training_plan_master to an Excele file
# with pd.ExcelWriter("training_plan_master.xlsx") as writer:
#     summary_df.to_excel(writer, sheet_name="Training Plan", index=False)

# workbook = load_workbook("training_plan_master.xlsx")
# sheet = workbook["Training Plan"]
# # Apply formatting
# for row in sheet.iter_rows(min_row=1, max_row=3):  # Format the header row
#     for cell in row:
#         cell.font = Font(bold=True)  # Set text to bold
#         cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Set background color to yellow

# workbook.save("training_plan_master.xlsx")
